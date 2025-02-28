import smtplib
import os
from email.message import EmailMessage
from openpyxl import load_workbook

def get_names_and_emails_from_excel(p_filename):
    wb = load_workbook(filename=p_filename)
    active = wb.active

    name_col = "A"
    email_col = "B"

    names = []
    emails = []
    for row in range(active.max_row):
        name = active[name_col][row].value
        email = active[email_col][row].value
        if (email != None):
            if (name == None):
                name = "Sir/Madam"
            names.append(name)
            emails.append(email)
    
    return names, emails

def get_email_data(p_filename):
    with open(p_filename, 'r') as f:
        subject = ""
        src_email = ""
        src_pass = ""
        salute = ""
        content = ""
        attachment_paths = []

        for line in f.readlines():
            if line.startswith('Source_addr: '):
                src_email = line[len('Source_addr: '):]
            if line.startswith('Password: '):
                src_pass = line[len('Password: '):]
            if line.startswith('Subject: '):
                subject = line[len('Subject: '):]
            if line.startswith('Salute: '):
                salute = line[len('Salute: '):]
            if line.startswith('Content: '):
                content = line[len('Content: '):]
            if line.startswith('Attachment_paths: '):
                attachment_paths = line[len('Attachment_paths: '):]

        src_email = src_email.split('\n')[0]
        src_pass = src_pass.split('\n')[0]
        subject = subject.split('\n')[0]
        salute = salute.split('\n')[0]
        content = content[:-1]
        attachment_paths = attachment_paths.split('\n')[0].replace(' ','').split(',')

    return subject, src_email, src_pass, salute, content, attachment_paths 

def append_name_to_content(p_dst_name, p_salute, p_content):
    if not p_salute.endswith(' '):
        p_salute = p_salute + " "
    updated_content = p_salute + p_dst_name + ",\n" + p_content
    return updated_content

names, emails = get_names_and_emails_from_excel('email_list.xlsx')
subject, src_email, src_pass, salute, content, attachment_paths = get_email_data("email_data.txt")

with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
    print("Server opened ...")
    server.login(src_email, src_pass)
    print("Logged in ...")

    for dst_name, dst_email in zip(names, emails):
        msg = EmailMessage()
        msg['Subject'] = subject
        msg['From'] = src_email
        msg['To'] = dst_email
        updated_content = append_name_to_content(dst_name, salute, content)
        msg.set_content(updated_content)

        for attc in attachment_paths:
            with open(attc, 'rb') as f:
                file_data = f.read()
                file_name = os.path.basename(attc)
                file_name, file_extension = os.path.splitext(file_name)
                file_extension = file_extension.replace('.','')
                msg.add_attachment(file_data, maintype = 'application', subtype = file_extension, filename = file_name)

        try:
            server.send_message(msg)
            print(f"Message sended to {dst_email}...")
        except:
            print(f"Could not send email to {dst_email}")
            pass
